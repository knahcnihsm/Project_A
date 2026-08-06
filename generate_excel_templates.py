"""
generate_excel_templates.py
============================
Generates two Excel workbooks for the Bulk Student Details Update feature:
  1. Bulk_Student_Update_Template.xlsx  – blank template (headers only)
  2. Bulk_Student_Update_Sample.xlsx    – sample template (headers + example rows)

Run:
    python generate_excel_templates.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Colour palette ────────────────────────────────────────────────────────────
NAVY        = "0D47A1"
NAVY_DARK   = "082F6A"
NAVY_LIGHT  = "E8EEF8"
WHITE       = "FFFFFF"
GRAY_LIGHT  = "F5F8FC"
GRAY_BORDER = "E6ECF5"

# ── Column definitions ────────────────────────────────────────────────────────
PERSONAL_COLS = [
    "Application No", "Register No", "Student Name", "Date of Birth",
    "Gender", "Aadhaar No", "Nationality", "District", "Caste / Category",
]
PARENT_COLS = [
    "Application No", "Register No", "Father Name", "Father Mobile",
    "Father Occupation", "Annual Family Income",
]
ADDRESS_COLS = [
    "Application No", "Register No", "Address Line", "PIN Code",
    "Landline No", "Mobile No", "Email ID",
]
ACADEMIC_COLS = [
    "Application No", "Register No", "Admission Category", "Program",
    "Department", "Batch", "Date of Admission",
]
QUALIFYING_COLS = [
    "Application No", "Register No", "Institution Name", "Institution Place",
    "Exam Passed", "Month & Year of Passing", "SSLC Percentage",
    "SSLC Register Number", "HSC Percentage", "HSC Register Number",
]
HSC_COLS = [
    "Application No", "Register No", "Stream", "Section", "Subject Name",
    "Month & Year", "Maximum Marks", "Marks Obtained",
]
DIPLOMA_COLS = [
    "Application No", "Register No", "Diploma Course", "Institution Name",
    "Board", "Second Year Percentage", "Third Year Percentage",
    "Aggregate Percentage",
]
PG_COLS = [
    "Application No", "Register No", "University Name", "University Place",
    "Institution Name", "Institution Place", "Exam Passed",
    "Month & Year of Passing", "Total Percentage",
    "Main Subject Percentage", "Degree Registration Number",
]

SHEETS = [
    ("Student Personal Details",        PERSONAL_COLS),
    ("Parent-Guardian Details",         PARENT_COLS),
    ("Communication Address",           ADDRESS_COLS),
    ("Permanent Address",               ADDRESS_COLS),
    ("Academic Admission Details",      ACADEMIC_COLS),
    ("Qualifying Exam (HSC-CBSE)",      QUALIFYING_COLS),
    ("HSC Marks",                       HSC_COLS),
    ("Diploma Qualification Details",   DIPLOMA_COLS),
    ("PG Qualifying Degree Details",    PG_COLS),
]

# ── Sample data rows ──────────────────────────────────────────────────────────
SAMPLE_ROWS = {
    "Student Personal Details": [
        ["RGCET/2026/9990", "26BTECH990", "Arun Kumar", "2007-05-14",
         "Male", "123456789012", "Indian", "Coimbatore", "BC"],
    ],
    "Parent-Guardian Details": [
        ["RGCET/2026/9990", "26BTECH990", "Kumarasamy", "9843012345",
         "Farmer", "250000"],
    ],
    "Communication Address": [
        ["RGCET/2026/9990", "26BTECH990",
         "12, Gandhi Street, Coimbatore", "641001",
         "04224231234", "9843012345", "arun.kumar@example.com"],
    ],
    "Permanent Address": [
        ["RGCET/2026/9990", "26BTECH990",
         "12, Gandhi Street, Coimbatore", "641001",
         "04224231234", "9843012345", "arun.kumar@example.com"],
    ],
    "Academic Admission Details": [
        ["RGCET/2026/9990", "26BTECH990", "Management",
         "First Year B.Tech", "Computer Science and Engineering",
         "2026-2030", "2026-08-01"],
    ],
    "Qualifying Exam (HSC-CBSE)": [
        ["RGCET/2026/9990", "26BTECH990",
         "Government Higher Secondary School", "Coimbatore",
         "HSC", "May 2025", "85.5", "SSLC2020123", "83.29", "HSC2020456"],
    ],
    "HSC Marks": [
        ["RGCET/2026/9990", "26BTECH990", "Science", "Academic",    "Mathematics",        "May 2025", 100, 92],
        ["RGCET/2026/9990", "26BTECH990", "Science", "Academic",    "Physics",            "May 2025", 100, 78],
        ["RGCET/2026/9990", "26BTECH990", "Science", "Academic",    "Chemistry",          "May 2025", 100, 81],
        ["RGCET/2026/9990", "26BTECH990", "Science", "Academic",    "Computer Science",   "May 2025", 100, 85],
        ["RGCET/2026/9990", "26BTECH990", "Science", "Vocational",  "Vocational Subject", "May 2025", 100, 74],
        ["RGCET/2026/9990", "26BTECH990", "Science", "Vocational",  "Related Subject I",  "May 2025", 100, 70],
        ["RGCET/2026/9990", "26BTECH990", "Science", "Vocational",  "Related Subject II", "May 2025", 100, 68],
    ],
    "Diploma Qualification Details": [
        ["RGCET/2026/9991", "26BTECH991",
         "Diploma in Mechanical Engineering", "PSG Polytechnic",
         "DOTE", "80", "82", "81.5"],
    ],
    "PG Qualifying Degree Details": [
        ["RGCET/2026/9992", "26PG9992", "Anna University", "Chennai",
         "RGCET", "Coimbatore", "B.E.", "May 2024",
         "75", "78", "PGREG001"],
    ],
}

# ── Instructions content ──────────────────────────────────────────────────────
INSTRUCTIONS = [
    ("BULK STUDENT DETAILS UPDATE — OFFICIAL TEMPLATE", "title"),
    ("", "gap"),
    ("1. PURPOSE", "section"),
    ("   Fill in the Excel sheets below to update existing student records in bulk.", "body"),
    ("   All sheets are committed together in a single transaction: if one row fails", "body"),
    ("   validation, NO rows are updated.", "body"),
    ("", "gap"),
    ("2. IDENTIFIERS", "section"),
    ("   Every data row must contain the student's Application No or Register No", "body"),
    ("   exactly as stored in the ERP. Students are matched against these values.", "body"),
    ("   Rows that match no existing student are reported as errors and block the commit.", "body"),
    ("", "gap"),
    ("3. BLANK CELLS = NO CHANGE", "section"),
    ("   Leave a cell empty to keep the existing value. Only filled-in cells are applied.", "body"),
    ("   Do NOT delete or rename the header row or the sheet names.", "body"),
    ("", "gap"),
    ("4. FILL ONLY WHAT YOU NEED", "section"),
    ("   You can update a single student, or thousands, in one workbook. Unused sheets", "body"),
    ("   may be left completely empty (only the header row).", "body"),
    ("", "gap"),
    ("5. VALIDATION RULES", "section"),
    ("   ERROR — blocks the entire commit:", "error"),
    ("     Gender: Male / Female / Transgender  (or M / F / T)", "body"),
    ("     Caste / Category: OC, BC, BCM, MBC, SC, SCA, ST  (OBC accepted as BC)", "body"),
    ("     Aadhaar No: exactly 12 digits", "body"),
    ("     Father Mobile / Mobile No: exactly 10 digits", "body"),
    ("     Email ID: must be a valid email address", "body"),
    ("     PIN Code: exactly 6 digits", "body"),
    ("     Percentages: between 0 and 100", "body"),
    ("     Marks: not negative and not more than Maximum Marks", "body"),
    ("     Admission Category / Program / Department must match master data", "body"),
    ("", "gap"),
    ("   WARNING — never blocks the commit:", "warning"),
    ("     Nationality, District, Batch, Board, institution names and Month & Year", "body"),
    ("     are free text — unusual values are accepted with a warning.", "body"),
    ("", "gap"),
    ("6. DERIVED VALUES", "section"),
    ("   Age, overall percentage, cut-off, diploma aggregate and fees are", "body"),
    ("   recalculated automatically — do NOT fill them.", "body"),
    ("", "gap"),
    ("7. SHEET 7 — HSC MARKS", "section"),
    ("   Enter ONE row per subject. Section must be 'Academic' or 'Vocational'.", "body"),
    ("   Academic stream: Mathematics, Physics, Chemistry and one more.", "body"),
    ("   Vocational stream: typically 3 subjects.", "body"),
    ("   Month & Year is optional for each subject row.", "body"),
    ("", "gap"),
    ("8. SHEETS OVERVIEW", "section"),
    ("     1   Student Personal Details", "body"),
    ("     2   Parent / Guardian Details", "body"),
    ("     3   Communication Address", "body"),
    ("     4   Permanent Address", "body"),
    ("     5   Academic Admission Details", "body"),
    ("     6   Qualifying Examination (HSC / CBSE)", "body"),
    ("     7   HSC Marks", "body"),
    ("     8   Diploma Qualification Details", "body"),
    ("     9   PG Qualifying Degree Details", "body"),
    ("    10   Instructions  (this sheet — documentation only, never parsed)", "body"),
    ("", "gap"),
    ("9. DATE FORMAT", "section"),
    ("   Use YYYY-MM-DD (e.g. 2007-05-14) or any common format like DD-MM-YYYY.", "body"),
]

SAMPLE_EXTRA = [
    ("", "gap"),
    ("10. SAMPLE DATA", "section"),
    ("    The other sheets contain example rows for a fictional student.", "body"),
    ("    Replace the values with real student data before uploading.", "body"),
    ("    Rows that do not match an existing student are rejected by the system.", "body"),
]

# ── Style helpers ─────────────────────────────────────────────────────────────

def _thin_border(color=GRAY_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _header_cell(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color=WHITE, name="Calibri", size=10)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border(NAVY_DARK)

def _data_cell(cell, value, row_idx):
    cell.value = value
    cell.font = Font(name="Calibri", size=10)
    bg = GRAY_LIGHT if row_idx % 2 == 0 else WHITE
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _thin_border(GRAY_BORDER)

def _setup_sheet(ws, columns):
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    for i, h in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 3)
    ws.sheet_properties.tabColor = NAVY

def _add_dropdowns(ws, columns, max_row=500):
    dv_map = {
        "Gender":          '"Male,Female,Transgender"',
        "Caste / Category":'"OC,BC,BCM,MBC,SC,SCA,ST"',
        "Section":         '"Academic,Vocational"',
        "Exam Passed":     '"HSC,CBSE,ICSE,Diploma,B.E.,B.Tech,B.Sc."',
    }
    for i, h in enumerate(columns, start=1):
        if h in dv_map:
            col = get_column_letter(i)
            dv = DataValidation(
                type="list", formula1=dv_map[h],
                showDropDown=False, showErrorMessage=True,
                errorTitle="Invalid value",
                error="Please select from the drop-down list.",
            )
            ws.add_data_validation(dv)
            dv.sqref = f"{col}2:{col}{max_row}"


# ── Workbook builder ──────────────────────────────────────────────────────────

def build_workbook(sample: bool) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Data sheets
    for sheet_name, columns in SHEETS:
        ws = wb.create_sheet(title=sheet_name)
        for ci, header in enumerate(columns, start=1):
            _header_cell(ws.cell(row=1, column=ci), header)
        if sample and sheet_name in SAMPLE_ROWS:
            for ri, row_data in enumerate(SAMPLE_ROWS[sheet_name], start=2):
                for ci, val in enumerate(row_data, start=1):
                    _data_cell(ws.cell(row=ri, column=ci), val, ri)
        _setup_sheet(ws, columns)
        _add_dropdowns(ws, columns)

    # Instructions sheet
    ws_i = wb.create_sheet(title="Instructions")
    ws_i.sheet_properties.tabColor = "607D8B"
    ws_i.column_dimensions["A"].width = 100

    FONT = {
        "title":   Font(name="Calibri", bold=True, size=14, color=NAVY_DARK),
        "section": Font(name="Calibri", bold=True, size=11, color=NAVY),
        "error":   Font(name="Calibri", bold=True, size=10, color="B71C1C"),
        "warning": Font(name="Calibri", bold=True, size=10, color="E65100"),
        "body":    Font(name="Calibri", size=10, color="333333"),
        "gap":     Font(name="Calibri", size=6),
    }
    FILL = {
        "title":   PatternFill("solid", fgColor=NAVY_LIGHT),
        "section": PatternFill("solid", fgColor="E3EAFC"),
        "error":   PatternFill("solid", fgColor="FFEBEE"),
        "warning": PatternFill("solid", fgColor="FFF3E0"),
        "body":    PatternFill("solid", fgColor=WHITE),
        "gap":     PatternFill("solid", fgColor=WHITE),
    }

    all_lines = list(INSTRUCTIONS)
    if sample:
        all_lines += SAMPLE_EXTRA

    for ri, (text, kind) in enumerate(all_lines, start=1):
        cell = ws_i.cell(row=ri, column=1, value=text)
        cell.font = FONT.get(kind, FONT["body"])
        cell.fill = FILL.get(kind, FILL["body"])
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws_i.row_dimensions[ri].height = 6 if kind == "gap" else 18

    ws_i.row_dimensions[1].height = 28

    return wb


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import os
    out_dir = os.path.dirname(os.path.abspath(__file__))

    blank_path  = os.path.join(out_dir, "Bulk_Student_Update_Template.xlsx")
    sample_path = os.path.join(out_dir, "Bulk_Student_Update_Sample.xlsx")

    print("Building Blank Template ...")
    wb_blank = build_workbook(sample=False)
    wb_blank.save(blank_path)
    print(f"  Saved -> {blank_path}")

    print("Building Sample Template ...")
    wb_sample = build_workbook(sample=True)
    wb_sample.save(sample_path)
    print(f"  Saved -> {sample_path}")

    print("\nDone! Both files are ready.")
